from datetime import datetime, timedelta
from typing import List
from tempfile import NamedTemporaryFile
import dateutil.parser
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

def generate_excel_document(issues: List, local_output_path: str, 
                            jql_query: str)  -> str:
    """This method will generate an Excel spreadsheet in .xlsx format (Open XML) which can then be uploaded into Tugboat Logic as an evidence file.
        
    Note that the generated output Excel document is in a specific format for Jira Server evidence. The rows and columns generated by this sample code are specific to Jira Server evidence, and as such, the would need to be altered for use with other integrations. 

    Returns the filepath for the Excel document that has been generated

    Params
    ------
    issues : List 
    The list of Jira issues that we would like to include within the Excel document
    local_output_path : str
    The local output path where we will save the Excel document for local review (optional)
    jql_query : str
    The Jira Query Language (JQL) query that was used to query these records from Jira Server"""
    # create the workbook
    wb = Workbook()
    # grab the active worksheet
    issues_ws = wb.active
    num_of_issues = len(issues)
    # setup the first worksheet for the list of Jira issues
    issues_ws = insert_jira_issues(issues_ws, issues)
    # set the number of columns required for the Jira issues worksheet
    issues_ws_cols = 10
    # setup the styling of the Jira worksheet
    issues_ws = default_worksheet_styles(ws=issues_ws, 
                                        title="Jira issues", 
                                        num_cols=issues_ws_cols, 
                                        num_rows=num_of_issues, 
                                        col_width=30, 
                                        row_height=50)                   
    # style the headers
    issues_ws = style_headers(ws=issues_ws,
                                header_cols=issues_ws_cols, 
                                header_rows=1, 
                                bold_header=True)
    # style the table body
    issues_ws = style_table_body(ws=issues_ws, min_row=2, 
                                 max_row=num_of_issues+2, min_col=1, 
                                 max_col=num_of_issues)
    # create the Report Details worksheet
    details_ws = wb.create_sheet("Report Details")
    # insert report details
    details_ws = insert_report_details(details_ws, jql_query)
    # set sheet layout
    details_ws_cols = 1 # 1 data column
    details_ws_rows = 6 # 6 data rows
    details_ws = default_worksheet_styles(ws=details_ws, 
                                            title="Repot Details", 
                                            num_cols=details_ws_cols, 
                                            num_rows=details_ws_rows, 
                                            col_width=30, 
                                            row_height=30)
    # style the headers
    details_ws = style_headers(ws=details_ws,
                                header_cols=details_ws_cols, 
                                header_rows=6, 
                                bold_header=False)
    # style the table body
    details_ws = style_table_body(ws=details_ws, min_row=1, 
                                  max_row=details_ws_rows, 
                                  min_col=details_ws_cols+1,
                                  max_col=details_ws_cols+1)
    # set Column B's width to 100
    details_ws.column_dimensions['B'].width = 100
    # Save the file locally if local output path is specified
    if local_output_path:
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        filename = f"{local_output_path}JiraServerEvidence_{timestamp}.xlsx"
        wb.save(filename) # save the file with the specified output path
    # delete=False forces the temporary file not to be deleted after context close
    with NamedTemporaryFile(suffix=".xlsx", mode='w+', delete=False) as tmp:
        tmp_file_path = tmp.name
        tmp.seek(0) # return cursor to the beginning of file
    # Save the file in a temporary file
    wb.save(tmp_file_path)
    # return the temporary file path
    return tmp_file_path
        

def default_worksheet_styles(ws: Worksheet, title: str, num_cols: int, 
                num_rows: int, col_width: int, row_height: int) -> Worksheet:
    """This helper method sets the default styles of the worksheet.
    
    Sets the sheet title, default column width and row height.

    Returns the styled worksheet

    Params
    ------
    ws : Worksheet
    The worksheet that is being styled.
    title : str
    The title of the worksheet.
    num_cols : int
    The number of data columns.
    num_rows: int
    The number of data rows.
    col_width : int
    The default width of the columns.
    row_height: int
    The default height of the rows."""
    # set the title of the worksheet
    ws.title = title
    # set default column widths of the worksheet
    # + 1 to compensate for 0 indexe drange
    for i in range(1,num_cols+1):
        ws.column_dimensions[get_column_letter(i)].width = col_width
    # set default row heights of the worksheet
    # + 2 to compensate for 0 indexed ranged and header row
    for i in range(num_rows+2):
        ws.row_dimensions[i].height = row_height
    return ws

def style_headers(ws: Worksheet,  header_cols: int, header_rows: int, 
                bold_header: bool) -> Worksheet:
    """This helper method sets the default styles of the headers in the worksheet.
    
    Sets the pattern fill, alignment and bold styles of the cells.

    Returns the styled worksheet

    Params
    ------
    ws : Worksheet
    The worksheet that is being styled.
    header_cols : int
    The number of header columns.
    header_rows : int
    The number of header rows.
    bold_header : bool
    Bold is added to the font if this is set to True."""
    # style the header cells
    for rows in ws.iter_rows(min_row=1, max_row=header_rows, 
                            min_col=1, max_col=header_cols):
        for cell in rows:
            cell.alignment = Alignment(vertical="bottom",wrap_text=True)
            cell.fill = PatternFill(patternType='gray125')
            if bold_header:
                cell.font = Font(bold=True)
    return ws

def style_table_body(ws: Worksheet, min_row: int, max_row: int,
                     min_col: int, max_col: int) -> Worksheet:
    """This helper method sets the default styles of the table body in the worksheet.
    
    Sets the alignment and text wrap of the cells.

    Returns the styled worksheet

    Params
    ------
    ws : Worksheet
    The worksheet that is being styled.
    min_row : int
    The starting row number of the body.
    max_row : int
    The max row number of the the body.
    min_row : int
    The starting column number of the body.
    max_row : int
    The max column number of the body."""
    # style the table body row by row, and cell by cell
    for rows in ws.iter_rows(min_row=min_row, max_row=max_row, 
                            min_col=min_col, max_col=max_col):
        for cell in rows:
            cell.alignment = Alignment(vertical="bottom",wrap_text=True)
    return ws

def insert_jira_issues(ws: Worksheet, issues: List) -> Worksheet:
    """This helper method transfers the list issues into the the worksheet.

    Returns the worksheet after data have been inserted.

    Params
    ------
    ws : Worksheet
    The worksheet that is being styled.
    issues : List
    The list of Jira issues to report to the Evidence Collector."""
    ws.append(["Issue number", "Issue type", "Project", "Summary", "Assignee",
                "Reporter", "Status", "Created (create date)", 
                "Resolved (resolution date)", "Issue URL"])
    
    for issue in issues:
        created_datetime = dateutil.parser.parse(issue.fields.created) \
                                          .strftime("%Y-%m-%d %H:%M:%S")
        resolution_datetime = dateutil.parser \
                                      .parse(issue.fields.resolutiondate) \
                                      .strftime("%Y-%m-%d %H:%M:%S")
        ws.append([issue.key, issue.fields.issuetype.name, 
                    issue.fields.project.name, issue.fields.summary, 
                    issue.fields.assignee.displayName, 
                    issue.fields.reporter.displayName, 
                    issue.fields.status.name, created_datetime, 
                    resolution_datetime, issue.self])
    return ws

def insert_report_details(ws: Worksheet, jql_query: str) -> Worksheet:
    """This helper that inserts the report details into the the worksheet.

    Returns the worksheet after report details have been inserted.

    Params
    ------
    ws : Worksheet
    The worksheet that is being styled.
    jql_query : str
    The query used in the Jira server request."""
    # current date time
    now = datetime.now()
    # append report detials
    ws.append(["Generated by", 
                    "Tugboat Logic® Jira Server Custom Evidence Collector"])
    ws.append(["Generated on", now.strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["Date range start",
                    (now - timedelta(days=90)).strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["Date range end", now.strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["JQL", jql_query])
    ws.append(["Fields requested",
                    ("issuetype, project, summary, assignee, reporter,"
                     " status, created, resolutiondate")])
    return ws